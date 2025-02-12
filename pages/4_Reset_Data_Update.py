from ctypes.wintypes import SIZE
import streamlit as st
import snowflake.connector
import snowflake
import Distro_Grid_Snowflake_Uploader
import datetime
import pandas as pd
from PIL import Image
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import openpyxl.utils.datetime as xl_datetime
import numpy as np
from io import BytesIO
from openpyxl import Workbook
import base64
from datetime import time
from Home import create_snowflake_connection
#from Home import log_connection_info
from Home import log_error_info
from Home import execute_query_and_close_connection
from ResetSH_formatter import format_RESET_schedule


from Reset_Schedule_to_Snowflake_Uploader import upload_reset_data
from openpyxl.utils.dataframe import dataframe_to_rows
import openpyxl
import datetime
from io import BytesIO
import numpy as np


    


#====================================================================================================================
# Setup page config and logo on the page
#====================================================================================================================

st.set_page_config(layout="wide", initial_sidebar_state="expanded")


def add_logo(logo_path, width, height):
    """Read and return a resized logo"""
    logo = Image.open(logo_path)
    modified_logo = logo.resize((width, height))
    return modified_logo

my_logo = add_logo(logo_path="./images/DeltaPacific_Logo.jpg", width=200, height = 100)
st.sidebar.image(my_logo)
st.sidebar.subheader("Delta Pacific Beverage Co.")




# Set Page Header   
st.header("CHAIN RESET MANAGMENT")
# Set custom CSS for hr element
st.markdown("""
        <style>
            hr {
                margin-top: 0.5rem;
                margin-bottom: 0.5rem;
                height: 3px;
                background-color: #333;
                border: none;
            }
        </style>
    """, unsafe_allow_html=True)

# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)

st.write("Download Template Here:")
# Add a hyperlink to your GitHub repository


# Your Streamlit app content

# Add a download link for the Excel file
st.markdown("[Download Reset Schedule Template Here:](https://github.com/rgriggs0072/ChainLinkAnalytics/raw/master/import_templates/RESET_SCHEDULE_TEMPLATE.xlsx)")

#===================================================================================================================

# Create a container to hold the file uploader
#===================================================================================================================
file_container = st.container()



#====================================================================================================================
# Function to retrieve Options_Name from the options_table to populate Chain Name dropdown 
#====================================================================================================================

# Function to retrieve options from Snowflake table
def get_options():
    # Create a connection
    conn, connection_id = create_snowflake_connection()

    try:
        # Execute the query and get the result using the function
        result = execute_query_and_close_connection('SELECT option_name FROM options_table ORDER BY option_name', conn, connection_id)

        # Extract options from the result
        options = [row[0] for row in result]

        return options

    except snowflake.connector.errors.Error as e:
        st.error(f"Error executing query: {str(e)}")
        log_error_info(str(e), connection_id)
        return None

    except Exception as e:
        st.error(f"An unexpected error occurred: {str(e)}")
        log_error_info(str(e), connection_id)
        return None

#====================================================================================================================
# END of Function to retrieve Options_Name from the options_table to populate Chain Name dropdown 
#====================================================================================================================

#--------------------------------------------------------------------------------------------------------------------

#====================================================================================================================
# Function that will update the options_table with a new chain option
#====================================================================================================================

# Function to update options in Snowflake table
def update_options(options):
    # Create a connection
    conn, connection_id = create_snowflake_connection()

    try:
        # Execute DELETE statement
        cursor = conn.cursor()
        cursor.execute('DELETE FROM options_table')

        # Execute INSERT statements for each option
        for option in options:
            cursor.execute("INSERT INTO options_table (option_name) VALUES (%s)", (option,))

        # Commit the changes
        conn.commit()

    except snowflake.connector.errors.Error as e:
        st.error(f"Error updating options: {str(e)}")
        log_error_info(str(e), connection_id)

    except Exception as e:
        st.error(f"An unexpected error occurred: {str(e)}")
        log_error_info(str(e), connection_id)

    finally:
        # Close the cursor and connection in the 'finally' block to ensure it happens even if an exception occurs
        cursor.close

#====================================================================================================================
# END of Function that will update the options_table with a new chain option
#====================================================================================================================

#====================================================================================================================
# Call the function Get_Options to populate session state and dropdown
#====================================================================================================================
# Retrieve options from Snowflake table
options = get_options()

# Initialize session state variables
if 'new_option' not in st.session_state:
    st.session_state.new_option = ""
if 'option_added' not in st.session_state:
    st.session_state.option_added = False

#====================================================================================================================
# End of  Call the function Get_Options to populate session state and dropdown
#====================================================================================================================

#--------------------------------------------------------------------------------------------------------------------

#====================================================================================================================
# Sets the subheader for the File Format uploader utility
#====================================================================================================================
with file_container:
    st.subheader(":blue[Reset Schedule File Format Utility]")
#====================================================================================================================
# END of Sets the subheader fro the File Format uploader utility
#====================================================================================================================

#--------------------------------------------------------------------------------------------------------------------

#===================================================================================================================
# Create the selection dropdown to select which Chain this section will format prior to upload to Snowflake
#===================================================================================================================
# Check if options are available
    if not options:
        st.warning("No options available. Please add options to the list.")
    else:
        # Create the dropdown in Streamlit
        selected_option = st.selectbox(':red[Select the Chain Reset Schedule to format]', options + ['Add new option...'], key="existing_option")

    # Check if the selected option is missing and allow the user to add it
    if selected_option == 'Add new option...':
        st.write("You selected: Add new option...")
        
        # Show the form to add a new option
        with st.form(key='add_option_form', clear_on_submit=True):
            new_option = st.text_input('Enter the new option', value=st.session_state.new_option)
            submit_button = st.form_submit_button('Add Option')
            
            if submit_button and new_option:
                options.append(new_option)
                update_options(options)
                st.success('Option added successfully!')
                st.session_state.option_added = True

        # Clear the text input field
        st.session_state.new_option = ""
        
    else:
        # Display the selected option
        #st.write(f":red[You selected {selected_option}]")

#===================================================================================================================
# END Option selection for sheet to format for which chain
#===================================================================================================================

#-------------------------------------------------------------------------------------------------------------------

#===================================================================================================================
# Creates the file uploader for the file to be formatted.  Also creates file to be downloaded and provides a button
# to download the file so it can be now uploaded to the uploader to import into snowflake
# ==================================================================================================================

        uploaded_file = st.file_uploader(":red[Upload reset schedule spreadsheet to be formatted]", type=["xlsx"])


        formatted_workbook = None  # Initialize the variable

    if st.button("Reformat Spreadsheet"):
        with st.spinner('Starting Format of Spreadsheet ...'):
            if uploaded_file is None:
                st.warning("Please upload a spreadsheet first.")
            else:
                # Load the workbook
                workbook = openpyxl.load_workbook(uploaded_file)

                formatted_workbook = format_RESET_schedule(workbook)
         
                # Call other formatting functions for different options
                formatted_workbook = workbook  # Use the original workbook    
                # Get sheet names from ExcelFile object
                sheet_names = workbook.get_sheet_names
                #st.write("Girl Friend, the sheet name is: ",sheet_names)

   
            # Create a new filename based on the selected option
            new_filename = f"formatted_{selected_option}_spreadsheet.xlsx"

            #Remove the autofilter if it exists
            if formatted_workbook is not None:
                for sheet_name in formatted_workbook.sheetnames:
                    sheet = formatted_workbook[sheet_name]
                    if sheet.auto_filter:
                        sheet.auto_filter.ref = None
                       

        # Check if the workbook was successfully formatted
        if formatted_workbook is not None:

            # Save the formatted workbook to a stream
            stream = BytesIO()
            formatted_workbook.save(stream)
            stream.seek(0)
    
           # st.write(formatted_workbook)
      
            # Provide the download link for the formatted spreadsheet
            st.download_button(
                label="Download formatted spreadsheet",
                data=stream.read(),
                file_name=new_filename,
                mime='application/vnd.ms-excel'
            )
        else:
            st.warning("No file has been prepared for download.")
        



#=========================================================================================================================================
# End of code bloack to Create the file uploader for the file to be formatted.  Also creates file to be downloaded and provides a button
# to download the file so it can be now uploaded to the uploader to import into snowflake
# ========================================================================================================================================

#-----------------------------------------------------------------------------------------------------------------------------------------


#=========================================================================================================================================
# Utility to process formatted Reset Schedule to Snowflake depending on which chain you are working on
#=========================================================================================================================================


with file_container:

    
        # Add horizontal line
    st.markdown("<hr>", unsafe_allow_html=True)
    st.subheader(":blue[Reset Schedule File to Upload to Snowflake Utility]")

    # Store the selected_option in session state
    if "selected_option" not in st.session_state:
        st.session_state.selected_option = None
    
    # Check if options are available
    if not options:
        st.warning("No options available. Please add options to the list.")
    else:
        # Create the dropdown in Streamlit
        selected_option = st.selectbox(':red[Select the Chain Reset Schedule to load to Snowflake]', options + ['Add new option...'], key="select_snowflake_option")

    # Check if the selected option is missing and allow the user to add it
    if selected_option == 'Add new option...':
        st.write("You selected: Add new option...")
        
        # Show the form to add a new option
        with st.form(key='add_option_form', clear_on_submit=True):
            new_option = st.text_input('Enter the new option', value=st.session_state.new_option)
            submit_button = st.form_submit_button('Add Option')
            
            if submit_button and new_option:
                options.append(new_option)
                update_options(options)
                st.success('Option added successfully!')
                st.session_state.option_added = True

        # Clear the text input field
        st.session_state.new_option = ""
        
    else:
        # Display the selected option
        #st.write(f"You selected: {selected_option}")

        # Store selected_option in session state
        st.session_state.selected_option = selected_option

#========================================================================================================================================
# This block of code creates uploader widget and prepares file to upload to snowflake once the file is uploaded it creates
# the button to start the upload process by the function: upload_reset_data in page Reset_Schedule_to_Snowflake_Uploader
#========================================================================================================================================

    # create file uploader
    uploaded_files = st.file_uploader(":red[Browse or select formatted reset schedule excel sheet To Upload to Snowflake]", type=["xlsx"], accept_multiple_files=True)

    # Process each uploaded file
    for uploaded_file in uploaded_files:
        # Read Excel file into pandas ExcelFile object
        excel_file = pd.ExcelFile(uploaded_file)
        #st.write("here is the file", excel_file)
        # Get sheet names from ExcelFile object
        sheet_names = excel_file.sheet_names
       
        # Display workbook name and sheet names in Streamlit
        workbook_name = uploaded_file.name
        
        ## Display DataFrame for each sheet in Streamlit
        for sheet_name in sheet_names:
            df = pd.read_excel(excel_file, sheet_name=sheet_name)

        #    # Modify DataFrame values directly to replace 'NAN' with empty string ''
            df = df.replace('NAN', np.nan)
            #st.write(df)

  
    # Check if any file has been uploaded
    if uploaded_files:
        # Write DataFrame to Snowflake on button click
        button_key = f"import_button_{workbook_name}_{sheet_name}"
        if st.button("Import into Snowflake", key=button_key):
            with st.spinner(f'Uploading {selected_option} data to Snowflake ...'):
                  upload_reset_data(df, selected_option)  
    else:
        st.warning("Please upload a file before attempting to import into Snowflake.")
    

    
#========================================================================================================================================
# END of block of code creates uploader widget and prepares file to upload to snowflake once the file is uploaded it creates
# the button to start the upload process by the function: upload_reset_data in page Reset_Schedule_to_Snowflake_Uploader
#========================================================================================================================================

#---------------------------------------------------------------------------------------------------------------------------------------

# =====================================================================================================================
# Query and Download Reset Schedule Data
# =====================================================================================================================

st.sidebar.markdown("<hr>", unsafe_allow_html=True)
st.sidebar.subheader(":blue[Download Reset Schedule Data]")

# Dropdown to select Chain Name
if not options:
    st.sidebar.warning("No chain options available. Please add options to the list.")
else:
    selected_chain = st.sidebar.selectbox("Select Chain Name for Reset Schedule", options, key="download_chain")

# Query data and download button
if selected_chain:
    if st.sidebar.button("Fetch Reset Schedule Data"):
        # Connect to Snowflake
        try:
            conn, connection_id = create_snowflake_connection()

            # Query data from the RESET_SCHEDULE_VIEW based on the selected chain
            query = f"""
                SELECT * 
                FROM RESET_SCHEDULE_VIEW
                WHERE CHAIN_NAME = %s;
            """
            cursor = conn.cursor()
            cursor.execute(query, (selected_chain,))
            data = cursor.fetchall()

            # Get column names
            columns = [col[0] for col in cursor.description]

            # Close the connection
            cursor.close()
            conn.close()

            # Convert data to a DataFrame
            if data:
                df = pd.DataFrame(data, columns=columns)

                # Create an Excel file in memory
                output = BytesIO()
                with pd.ExcelWriter(output, engine='openpyxl') as writer:
                    df.to_excel(writer, index=False, sheet_name="Reset_Schedule")
                output.seek(0)

                # Provide a download button
                st.sidebar.download_button(
                    label=f"Download {selected_chain} Reset Schedule",
                    data=output,
                    file_name=f"{selected_chain}_Reset_Schedule.xlsx",
                    mime="application/vnd.ms-excel"
                )
            else:
                st.sidebar.warning(f"No data found for chain: {selected_chain}")

        except Exception as e:
            st.sidebar.error(f"An error occurred: {str(e)}")

# =====================================================================================================================
# END of Query and Download Reset Schedule Data
# =====================================================================================================================
