import streamlit as st
import pandas as pd
import openpyxl
from PIL import Image
import snowflake.connector
from openpyxl import load_workbook
import streamlit.components.v1 as components
from datetime import datetime
import logging
import uuid
from io import BytesIO
from streamlit.elements.image import MAXIMUM_CONTENT_WIDTH

# ===================== Custom imports ============================================================================== 

from DG_Non_Pivot_Format import format_non_pivot_table
from misc_distro_grid_uploader import update_spinner, upload_distro_grid_to_snowflake



# ===================== End Custom Imports ==========================================================================





# ==================================================================================================================
# THIS SECTION OF CODE HANDLES THE LOGO AND SETS THE VIEW TO WIDE
# ==================================================================================================================

def add_logo(logo_path, width, height):
    """Read and return a resized logo"""
    logo = Image.open(logo_path)
    modified_logo = logo.resize((width, height))
    return modified_logo

my_logo = add_logo(logo_path="./images/DeltaPacific_Logo.jpg", width=200, height=100)
st.sidebar.image(my_logo)
st.sidebar.subheader("Delta Pacific Beverage Co.")
st.subheader("Misc Distribution Grid Processing")
st.text("This processing will add items to the Distribution Grid table for a particular chain where items need to be added rather than a rebuilt:")

## Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)
st.write("Download Template Here:")
# Add a hyperlink to your GitHub repository

# Add a download link for the Excel file templates
st.markdown(
    "[Download Distro Grid Template](https://github.com/rgriggs0072/ChainLinkAnalytics/raw/master/import_templates/Distribution_Grid_Template.xlsx)")




# ============================================================================================================================================================
# 11/28/2023 - Randy Griggs - Function to create connection to the database
# ============================================================================================================================================================

# Function to create and return a Snowflake connection object with logging
def create_snowflake_connection():
    try:
        # Load Snowflake credentials from the secrets.toml file
        snowflake_creds = st.secrets["snowflake"]

        # Create a connection ID
        connection_id = str(uuid.uuid4())

        # Create and return a Snowflake connection object
        conn = snowflake.connector.connect(
            account=snowflake_creds["account"],
            user=snowflake_creds["user"],
            password=snowflake_creds["password"],
            warehouse=snowflake_creds["warehouse"],
            database=snowflake_creds["database"],
            schema=snowflake_creds["schema"]
        )

        return conn, connection_id

    except snowflake.connector.errors.Error as e:
        st.error(f"Error creating Snowflake connection: {str(e)}")
        # Log the error or take appropriate action
        log_error_info(str(e), connection_id)
        return None, None  # Return None to indicate an error


# ============================================================================================================================================================
# END of Function 11/28/2023 - Randy Griggs - Function to create connection to the database
# ============================================================================================================================================================

# ==================================================================================================================
# HERE WE CREATE THE FUNCTION TO GET THE CHAIN OPTIONS FROM SNOWFLAKE FOR THE DROPDOWN
# ==================================================================================================================

def get_options():
    conn = create_snowflake_connection()[0]
    cursor = conn.cursor()
    cursor.execute('SELECT option_name FROM options_table ORDER BY option_name')
    options = [row[0] for row in cursor]
    return options

# ==================================================================================================================
# FUNCTION TO UPDATE THE CHAIN OPTIONS IN SNOWFLAKE FOR THE DROPDOWN
# ==================================================================================================================

def update_options(options):
    conn = create_snowflake_connection()[0]
    cursor = conn.cursor()
    cursor.execute('DELETE FROM options_table')
    for option in options:
        cursor.execute("INSERT INTO options_table (option_name) VALUES (%s)", (option,))
    conn.commit()

# ==================================================================================================================
# Create a container to hold the file uploader and the dropdown for selecting the chain distro grid.
# ==================================================================================================================

file_container = st.container()
with file_container:
    st.subheader(":blue[Distro Grid Fomatting Utility]")

options = get_options()

# Initialize session state variables
if 'new_option' not in st.session_state:
    st.session_state.new_option = ""
if 'option_added' not in st.session_state:
    st.session_state.option_added = False

# Create the dropdown and allow adding new options if needed
if not options:
    st.warning("No options available. Please add options to the list.")
else:
    selected_option = st.selectbox(':red[Select the Chain Distro Grid to format]', options + ['Add new option...'], key="existing_option")
    st.session_state.selected_option = selected_option

if selected_option == 'Add new option...':
    new_option = st.text_input('Enter the new option', value=st.session_state.new_option)
    if st.form_submit_button('Add Option') and new_option:
        options.append(new_option)
        update_options(options)
        st.success('Option added successfully!')
        st.session_state.new_option = ""
else:
    st.write(f"You selected: {selected_option}")

# ======================================================================================================================
# USING THE CONTAINER CREATED EARLIER, CREATES THE UPLOADER FOR FORMATTING THE DISTRIBUTION GRID SPREADSHEET.
# Handle formatting based on whether the spreadsheet is a pivot table or not.
# ======================================================================================================================

# Initialize session state for warnings acknowledgment and file handling
if 'acknowledged_warnings' not in st.session_state:
    st.session_state['acknowledged_warnings'] = False
if 'warnings_present' not in st.session_state:
    st.session_state['warnings_present'] = False
if 'file_uploaded' not in st.session_state:
    st.session_state['file_uploaded'] = False

uploaded_file = st.file_uploader(":red[Browse or drag here the Distribution Grid to Format]", type=["xlsx"], key="uploaded_file")

# If a file is uploaded
if uploaded_file:
    st.session_state['file_uploaded'] = True
    workbook = openpyxl.load_workbook(uploaded_file)

    # If the "Reformat DG Spreadsheet" button is clicked
    if st.button("Reformat DG Spreadsheet"):
        formatted_workbook = None
        stream = BytesIO()

        # Format the non-pivot table
        formatted_workbook = format_non_pivot_table(workbook, stream, st.session_state.selected_option)

        # Validate chain_name in the uploaded file against the selected option
        df = pd.read_excel(uploaded_file, sheet_name=workbook.sheetnames[0])
        df.columns = df.columns.str.strip()  # Strip whitespace from column names

        if 'CHAIN_NAME' in df.columns:
            file_chain_name = df['CHAIN_NAME'].iloc[0].strip()  # Ensure there are no leading/trailing spaces
            if file_chain_name.lower() != st.session_state.selected_option.lower():
                st.error(f"Chain name in the file ({file_chain_name}) does not match the selected option ({st.session_state.selected_option}). Please correct the file or select the correct chain.")
                st.stop()  # Stop execution to prevent further processing
        else:
            st.error(f"'CHAIN_NAME' column not found in the uploaded file ({uploaded_file.name}). Please make sure the file has the correct format.")
            st.stop()  # Stop execution to prevent further processing

        # If warnings are present and have not been acknowledged
        if st.session_state['warnings_present'] and not st.session_state['acknowledged_warnings']:
            st.warning("Warnings detected! Please remove file, correct the issues listed, and re-upload the file.")
        else:
            st.session_state['acknowledged_warnings'] = True  # Mark as acknowledged

            # Allow file download if warnings were acknowledged or no warnings were present
            if formatted_workbook and st.session_state['acknowledged_warnings']:
                formatted_workbook.save(stream)
                stream.seek(0)
                st.download_button(
                    label="Download formatted spreadsheet",
                    data=stream,
                    file_name="formatted_spreadsheet.xlsx",
                    mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
                )


# ===========================================================================================================
# create code uploader in preparation to write to snowflake
# ==========================================================================================================

# Create a container to hold the file uploader
snowflake_file_container = st.container()

# Add a title to the container
with snowflake_file_container:
    st.markdown("<hr>", unsafe_allow_html=True)
    st.subheader(":blue[Write Distribution Grid to Snowflake Utility]")

with snowflake_file_container:
    conn = create_snowflake_connection()[0]  # Get connection object

# Initialize session state variables
if 'new_option' not in st.session_state:
    st.session_state.new_option = ""
if 'option_added' not in st.session_state:
    st.session_state.option_added = False

# Check if options are available
if not options:
    st.warning("No options available. Please add options to the list.")
else:
    # Create the dropdown in Streamlit
    selected_option = st.selectbox(':red[Select the Chain Distro Grid to upload to Snowflake]',
                                   options + ['Add new option...'], key="existing_chain_option")

    # Check if the selected option is missing and allow the user to add it
    if selected_option == 'Add new option...':
        st.write("You selected: Add new option...")

        # Show the form to add a new option
        with st.form(key='add_option_form', clear_on_submit=True):
            new_option = st.text_input('Enter the new option', value=st.session_state.new_option)
            submit_button = st.form_submit_button('Add Option')

            if submit_button and new_option:
                options.append(new_option)
                update_options(options)
                st.success('Option added successfully!')
                st.session_state.option_added = True

        # Clear the text input field
        st.session_state.new_option = ""

    else:
        # Display the selected option
        st.write(f"You selected: {selected_option}")

# Create file uploader
uploaded_files = st.file_uploader("Browse or select formatted Distribution Grid excel sheets", type=["xlsx"],
                                  accept_multiple_files=True)

# Process each uploaded file
for uploaded_file in uploaded_files:
    # Read Excel file into pandas ExcelFile object
    excel_file = pd.ExcelFile(uploaded_file)

    # Get sheet names from ExcelFile object
    sheet_names = excel_file.sheet_names

    # Process each sheet in the Excel file
    for sheet_name in sheet_names:
        df = pd.read_excel(excel_file, sheet_name=sheet_name)

        # Standardize column names to remove leading/trailing whitespace
        df.columns = df.columns.str.strip()

        # Check if the 'MANUFACTURER' column has any empty cells
        if 'MANUFACTURER' in df.columns:
            if df['MANUFACTURER'].isnull().any():
                st.error(f"The 'MANUFACTURER' column in the uploaded spreadsheet '{uploaded_file.name}', sheet '{sheet_name}', contains empty cells. Please correct these before proceeding.")
                continue  # Skip further processing for this file
        else:
            st.error(f"'MANUFACTURER' column not found in the uploaded file ({uploaded_file.name}). Please make sure the file has the correct format.")
            continue

        # Validate chain_name in the uploaded file against the selected option to ensure they match
        if 'CHAIN_NAME' in df.columns:
            file_chain_name = df['CHAIN_NAME'].iloc[0].strip()  # Ensure there are no leading/trailing spaces
            if file_chain_name.lower() != selected_option.lower():
                st.error(f"Chain name in the file ({file_chain_name}) does not match the selected option ({selected_option}). Please correct the file or select the correct chain.")
                continue  # Skip further processing for this file if chain names do not match
        else:
            st.error(f"'CHAIN_NAME' column not found in the uploaded file ({uploaded_file.name}). Please make sure the file has the correct format.")
            continue

        # Fetch existing data for the selected chain from the distro_grid table
        query = f"SELECT * FROM distro_grid WHERE chain_name = '{selected_option}'"
        existing_data = pd.read_sql(query, conn)

        # Check if existing data is not empty
        if not existing_data.empty:
            # Compare 'MANUFACTURER' column in both dataframes
            if 'MANUFACTURER' in existing_data.columns:
                new_manufacturer_values = set(df['MANUFACTURER'].str.strip().unique())
                existing_manufacturer_values = set(existing_data['MANUFACTURER'].str.strip().unique())

                # If there is overlap in manufacturer values, warn the user
                if new_manufacturer_values & existing_manufacturer_values:
                    st.warning(f"The new data for chain '{selected_option}' contains manufacturers already present in the existing distro_grid table. Please upload the alcohol distro grid spreadsheet first before the misc data.")
                    continue  # Skip further processing for this file

        # Display DataFrame for each sheet in Streamlit
        st.write(f"Sheet: {sheet_name} and Chain Selected is: {selected_option}")
        # st.dataframe(df)


        # Write DataFrame to Snowflake on button click
        button_key = f"import_button_{uploaded_file.name}_{sheet_name}"
        if st.button("Import Distro Grid into Snowflake", key=button_key):
            with st.spinner('Uploading data to Snowflake ...'):
                # Write DataFrame to Snowflake based on the selected store
                upload_distro_grid_to_snowflake(df, selected_option, update_spinner)

 # ===========================================================================================================
# Create a button to download the existing distro_grid data for the selected chain
# ===========================================================================================================

# Create a button to download the existing distro_grid data for the selected chain
if st.button("Download Existing Distro Grid Data for Selected Chain"):
    # Fetch existing data for the selected chain from the distro_grid table
    conn = create_snowflake_connection()[0]
    query = f"SELECT * FROM distro_grid WHERE chain_name = '{selected_option}'"
    existing_data = pd.read_sql(query, conn)

    if existing_data.empty:
        st.warning(f"No existing data found for chain '{selected_option}'.")
    else:
        # Convert the existing data to an Excel file and provide a download link
        download_stream = BytesIO()
        with pd.ExcelWriter(download_stream, engine='openpyxl') as writer:
            existing_data.to_excel(writer, sheet_name='Distro Grid Data', index=False)
        download_stream.seek(0)

        st.download_button(
            label=f"Download Distro Grid Data for {selected_option}",
            data=download_stream,
            file_name=f"{selected_option}_Distro_Grid.xlsx",
            mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

