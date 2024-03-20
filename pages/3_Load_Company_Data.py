# Import required libraries
from importlib.resources import Package
from multiprocessing.connection import wait
import snowflake.connector
import streamlit as st
import pandas as pd
import openpyxl
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook
import numpy as np
from io import BytesIO
from Home import create_snowflake_connection
#from Home import log_connection_info
from Home import log_error_info
from openpyxl import Workbook
from datetime import datetime
from PIL import Image
import streamlit.components.v1 as components



#=====================================================================================================================================
#Page configuration and page formatting
#=====================================================================================================================================

st.set_page_config(layout="wide")
padding_top = 10

#reduces the space above the header so the header will be at the top of the page
reduce_header_height_style = """
    <style>
        div.block-container {padding-top:1rem;}
    </style>
"""
st.markdown(reduce_header_height_style, unsafe_allow_html=True)





def add_logo(logo_path, width, height):
    """Read and return a resized logo"""
    logo = Image.open(logo_path)
    modified_logo = logo.resize((width, height))
    return modified_logo

my_logo = add_logo(logo_path="./images/DeltaPacific_Logo.jpg", width=200, height = 100)
st.sidebar.image(my_logo)
st.sidebar.subheader("Delta Pacific Beverage Co.")



# Set Page Header
st.header("Load Customer, Products and Supplier by County Data")

# Set custom CSS for hr element
st.markdown("""
        <style>
            hr {
                margin-top: 0.5rem;
                margin-bottom: 0.5rem;
                height: 3px;
                background-color: #333;
                border: none;
            }
        </style>
    """, unsafe_allow_html=True)

# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)

#=====================================================================================================================================
# End Page configuration and page formatting
#=====================================================================================================================================


#-------------------------------------------------------------------------------------------------------------------------------------

#=====================================================================================================================================
# The following function will reformat the Customers Excel worksheet into importable Data
#=====================================================================================================================================
def format_Customers_Dataload(workbook):
    # Delete all sheets except Customers
    for sheet_name in workbook.sheetnames:
        if sheet_name != 'Customers':
            workbook.remove(workbook[sheet_name])

    # Select the Customers sheet
    ws = workbook['Customers']

    # Remove filters
    ws.auto_filter.ref = None

    # Create a new column for store name
    ws.insert_cols(3)
    #ws.cell(row=1, column=3, value='STORE NAME')
    
    # Copy values before the # to store name column
    for row in ws.iter_rows(min_row=2, min_col=4, max_col=4):
        for cell in row:
            if '#' in str(cell.value):
                cell_offset = ws.cell(row=cell.row, column=4)
                store_name = str(cell.value).split('#')[0].replace("#", "")
                cell_offset.value = store_name



    # Move values from column D to column C
    for row in ws.iter_rows(min_row=2, min_col=5, max_col=5):
        for cell in row:
            cell_offset = ws.cell(row=cell.row, column=3)
            cell_offset.value = cell.value

    # Delete original column D (now column E)
    ws.delete_cols(5)
   
    #ws.cell(row=1, column=3, value='Store_Number')


    # Rename Columns as required to meet objective for uploading to Snowflake
    ws.cell(row=1, column=1, value='Customer_id')
    ws.cell(row=1, column=2, value='Chain_Name')
    ws.cell(row=1, column=3, value='Store_Number')
    ws.cell(row=1, column=4, value='Store_Name')
    ws.cell(row=1, column=5, value='Address')
    ws.cell(row=1, column=6, value='City')
    ws.cell(row=1, column=7, value='County')
    ws.cell(row=1, column=8, value='Salesperson')
    ws.cell(row=1, column=9, value='Account_Status')
    


    # Remove all apostrophes (') from column E
    for cell in ws['E']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace("'", "")

    # Remove all apostrophes (') from column B 
    for cell in ws['B']:
        if cell.value is not None and isinstance(cell.value, str):
            cell.value = cell.value.replace("'", "")



   



    return workbook

#=====================================================================================================================================
# END The following function will reformat the Customers Excel worksheet into importable Data
#=====================================================================================================================================

file_container = st.container()
#-------------------------------------------------------------------------------------------------------------------------------------

#=====================================================================================================================================
# Handles page look, creates uploader and once uploaded provides button to call the format function
#=====================================================================================================================================
with file_container:
    #st.markdown("<hr>", unsafe_allow_html=True)
    st.subheader(":blue[Process Customer Formatting and Uploading to Snowflake]")
# Make the text red
st.markdown('<p style="color: red;">Upload Customer Report from Encompass</p>', unsafe_allow_html=True)

# Upload the workbook
uploaded_file = st.file_uploader("Browse or Select File", type=["xlsx", "xls"])

if uploaded_file is not None:
    # Load the workbook
    workbook = openpyxl.load_workbook(uploaded_file)

    # Show the Reformat button upon selecting will call the format_Customers_dataload function
    if st.button("Reformat"):
        # Format the sales report
        new_workbook = format_Customers_Dataload(workbook)

        # Download the formatted file
        new_filename = 'formatted_' + uploaded_file.name
        stream = BytesIO()
        new_workbook.save(stream)
        stream.seek(0)
        st.download_button(label="Download formatted file", data=stream.read(), file_name=new_filename, mime='application/vnd.ms-excel')
        
#=====================================================================================================================================
# END Handles page look, creates uploader and once uploaded provides button to call the format function
#=====================================================================================================================================

#-----------------------------------------------------------------------------------------------------------------------------------------
    
#=====================================================================================================================================
# The below function formats the Products report to prepare for upload to snowflake table
#=====================================================================================================================================


def format_Products_Dataload(workbook):
    # Select the Products sheet
    ws = workbook['Products']

    # Get the column data from Column G
    column_g_data = [cell.value for cell in ws['G']]

    # Delete data in Column G
    for cell in ws['G']:
        cell.value = None

    # Insert a new column at Column B
    ws.insert_cols(2)

    # Populate new Column B with data from Column G
    for cell, value in zip(ws['B'], column_g_data):
        cell.value = value
        

    # Get the column data from Column E
    column_g_data = [cell.value for cell in ws['E']]

    # Delete data in Column 
    # for cell in ws['G']:
    #     cell.value = None

    # # Insert a new column at Column B
    # ws.insert_cols(4)

    # Populate new Column D with data from Column E
    for cell, value in zip(ws['D'], column_g_data):
        cell.value = value
       
    # Delete columns E 
    ws.delete_cols(ws['E'][0].column, amount=1)  # Delete column E
    
    # Remove hyphens from values in Column E
    for cell in ws['E']:
        if isinstance(cell.value, str) and cell.value is not None:
            cell.value = cell.value.replace('-', '')

    # Get the last row with data
    last_row = ws.max_row

    # Get the last column with data
    last_col = ws.max_column

    # Remove commas from cell values in specified columns
    for row_number in range(2, last_row + 1):
        for col_number in range(1, last_col + 1):
            cell_value = ws.cell(row=row_number, column=col_number).value
            if isinstance(cell_value, str):
                cell = ws.cell(row=row_number, column=col_number)
                cell.value = cell_value.replace(',', '')
                
        # Remove commas from cell values in specified columns
    for row_number in range(2, last_row + 1):
        for col_number in range(1, last_col + 1):
            cell_value = ws.cell(row=row_number, column=col_number).value
            if isinstance(cell_value, str):
                cell = ws.cell(row=row_number, column=col_number)
                cell.value = cell_value.replace("'", '')


                

    # Rename column F to PRODUCT_MANAGER
    ws['F1'].value = 'PRODUCT_MANAGER'

    # Clear data in the PRODUCT_MANAGER column
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=6, max_col=6):
        for cell in row:
            cell.value = None

       
    # Delete columns G 
    ws.delete_cols(ws['G'][0].column, amount=1)  # Delete column G
    
    # Find the column index for the column you want to fill
    column_index = 5  # Assuming column E is the Carrier_upc column

    # Fill empty cells with the value 1 in the carrier_upc column
    for row_number in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_number, column=column_index)
        if cell.value is None:
            cell.value = 999999999999
    
    return workbook


#=====================================================================================================================================
# END The below function formats the Products report to prepare for upload to snowflake table
#=====================================================================================================================================

#------------------------------------------------------------------------------------------------------------------------------------




#-------------------------------------------------------------------------------------------------------------------------------------

#=====================================================================================================================================
# The below function is called when a new product table needs to be build with the formatted spreasheet from the clustomers products 
# list
#=====================================================================================================================================

def write_to_products_snowflake(df):
    
    
    # read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    
    
    
    
    # Modify DataFrame values directly to replace 'NAN' with empty string ''
    
    df = df.replace('NAN', np.nan)
    
    # Replace NaN values with NULL in the DataFrame
    df.fillna(value='NULL', inplace=True)
    
    # Create a connection using function create_snowflake_connection()
    conn, connection_id = create_snowflake_connection()
    
    
    
    # Convert the DataFrame to a list of tuples
    data_tuples = [tuple(row) for row in df.values]

    # Get the current timestamp
    current_timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
   
    
   # write DataFrame to Snowflake
    cursor = conn.cursor()
    
    # Generate the SQL query
    values = ', '.join([str(tuple(row)) for row in df.values])
    sql_query = f"""
    CREATE OR REPLACE TABLE PRODUCTS AS
    SELECT
        CAST(PRODUCT_ID AS NUMBER) AS PRODUCT_ID,
        CAST(SUPPLIER AS VARCHAR) AS SUPPLIER,
        CAST(PRODUCT_NAME AS VARCHAR) AS PRODUCT_NAME,
        CAST(PACKAGE AS VARCHAR) AS PACKAGE,
        CAST(CARRIER_UPC AS NUMBER) AS CARRIER_UPC,
        CAST(PRODUCT_MANAGER AS VARCHAR) AS PRODUCT_MANAGER,
        CAST('{current_timestamp}' AS TIMESTAMP) AS UPLOAD_DATE
    FROM
        (VALUES {values}) AS tmp(
            PRODUCT_ID,
            SUPPLIER,
            PRODUCT_NAME,
            PACKAGE,
            CARRIER_UPC,
            PRODUCT_MANAGER
        );
    """

    #st.write (sql_query)
     # Execute the SQL query
    cursor.execute(sql_query)

    # Commit the changes to the Snowflake table
    conn.commit()

    # Close the cursor and the connection
    cursor.close()
    conn.close()

#=====================================================================================================================================
# END The below function is called when a new product table needs to be build with the formatted spreasheet from the clustomers products 
# list
#=====================================================================================================================================

#-------------------------------------------------------------------------------------------------------------------------------------

#=====================================================================================================================================
# Function that is called to write Customer table data to the customers  table in snowflake
#=====================================================================================================================================
def write_to_customers_snowflake(df):
    # # replace NaN values with "NULL"
    df.fillna(value="NULL", inplace=True)

    # Create a connection using function create_snowflake_connection()
    conn, connection_id = create_snowflake_connection()
    
    
    # write DataFrame to Snowflake
    cursor = conn.cursor()

    # Apply TRIM and UPPER to each column in the DataFrame
    df = df.applymap(lambda x: str(x).strip().upper() if pd.notna(x) else x)

    # Generate the SQL query
    values = ', '.join([str(tuple(row)) for row in df.values])
    sql_query = f"""
    CREATE OR REPLACE TABLE CUSTOMERS AS
    SELECT
        CAST(CUSTOMER_ID AS NUMBER) AS CUSTOMER_ID,
        CAST(TRIM(UPPER(CHAIN_NAME)) AS VARCHAR) AS CHAIN_NAME,
        CAST(STORE_NUMBER AS NUMBER) AS STORE_NUMBER,
        CAST(TRIM(UPPER(STORE_NAME)) AS VARCHAR) AS STORE_NAME,
        CAST(TRIM(UPPER(ADDRESS)) AS VARCHAR) AS ADDRESS,
        CAST(TRIM(UPPER(CITY)) AS VARCHAR) AS CITY,
        CAST(TRIM(UPPER(COUNTY)) AS VARCHAR) AS COUNTY,
        CAST(TRIM(UPPER(SALESPERSON)) AS VARCHAR) AS SALESPERSON,
        CAST(TRIM(UPPER(ACCOUNT_STATUS)) AS VARCHAR) AS ACCOUNT_STATUS
    FROM
        (VALUES {values}) AS tmp(
            CUSTOMER_ID,
            CHAIN_NAME,
            STORE_NUMBER,
            STORE_NAME,
            ADDRESS,
            CITY,
            COUNTY,
            SALESPERSON,
            ACCOUNT_STATUS
        );
    """

    # Execute the SQL query
    cursor.execute(sql_query)

    # Commit the changes to the Snowflake table
    conn.commit()

    # Close the cursor and the connection
    cursor.close()
    conn.close()

#=====================================================================================================================================
# END Function that is called to write Customer table data to the customers  table in snowflake
#=====================================================================================================================================

#-------------------------------------------------------------------------------------------------------------------------------------



#-------------------------------------------------------------------------------------------------------------------------------------

#======================================================================================================================================
# The below block of code handles the upload widget and then calls the write_to_customers_snowflake function to write the data
# to snowflake
#======================================================================================================================================

# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)
# Make the text red
st.markdown('<p style="color: red;">Upload Customer Report After it has Been Formatted</p>', unsafe_allow_html=True)

# create file uploader
uploaded_file = st.file_uploader("Browse or Select File1", type=["xlsx"])

# check if file was uploaded
if uploaded_file:
    # read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    # display DataFrame in Streamlit
    #st.dataframe(df)

    # write DataFrame Customers to Snowflake Customers Table on button click
    if st.button("Import into Snowflake"):
        with st.spinner('Uploading Customers data to Snowflake ...'):
            write_to_customers_snowflake(df)
            st.write("Customer Data has been imported into Snowflake table! ", "Customers")

#======================================================================================================================================
# END The below block of code handles the upload widget and then calls the write_to_customers_snowflake function to write the data
# to snowflake
#======================================================================================================================================

#--------------------------------------------------------------------------------------------------------------------------------------

#=====================================================================================================================================
# The below block of code provides the upload and reformat button for the products report from company's products list tables
# and calls the function format_Products_Dataload to do the formatting
#=====================================================================================================================================




#with file_container:
st.markdown("<hr>", unsafe_allow_html=True)
st.subheader(":blue[Process Products Formatting and Uploading to Snowflake]")
# Make the text red
st.markdown('<p style="color: red;">Upload Products Report from Encompass</p>', unsafe_allow_html=True)


# Upload the workbook
uploaded_file = st.file_uploader("Browse or select File", type=["xlsx", "xls"])

if uploaded_file is not None:
    # Load the workbook
    workbook = openpyxl.load_workbook(uploaded_file)

    # Show the Reformat button
    if st.button("Reformat_Products"):
        # Format the sales report
        new_workbook = format_Products_Dataload(workbook)

        # Download the formatted file
        new_filename = 'formatted_' + uploaded_file.name
        stream = BytesIO()
        new_workbook.save(stream)
        stream.seek(0)
        st.download_button(label="Download formatted file", data=stream.read(), file_name=new_filename, mime='application/vnd.ms-excel')
    

#=====================================================================================================================================
# END The below block of code provides the upload and reformat button for the products report from company's products list tables
# and calls the function format_Products_Dataload to do the formatting
#=====================================================================================================================================r

#=====================================================================================================================================
# The below block of code handles the upload of the formatted products spreasheet and then calls the function 
# write_to_products_snowflake to upload the data to the products table in snowflake
#=====================================================================================================================================

# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)
# Make the text red
st.markdown('<p style="color: red;">Upload Products Table After it has been formatted</p>', unsafe_allow_html=True)

# create file uploader
uploaded_file = st.file_uploader("Browse or Select File", type=["xlsx"])


# check if file was uploaded
if uploaded_file:
    # read Excel file into pandas DataFrame
    df = pd.read_excel(uploaded_file)
    #print(df.columns)
    # display DataFrame in Streamlit
    #st.dataframe(df)


    # write DataFrame to Snowflake on button click
    if st.button("Import into Snowflake", key="import_button"):
        with st.spinner('Uploading product data to Snowflake ...'):
            write_to_products_snowflake(df)
            st.write("Product Data has been imported into Snowflake table! ", "Products")




#=====================================================================================================================================
# END The below block of code handles the upload of the formatted products spreasheet and then calls the function 
# write_to_products_snowflake to upload the data to the products table in snowflake
#=====================================================================================================================================

# Create a container to hold the file uploader
#===================================================================================================================
file_container = st.container()


#=======================================================================================================================
# This function takes the Supplier by county pivot table and formats it so it can be uploaded to snowflake
#=======================================================================================================================
#
def format_supplier_by_county(file_content):
    df_formatted = None

    # Load the Excel file
    xls = pd.ExcelFile(file_content)

    # Check if the expected sheet ("Report") is present
    if "Report" not in xls.sheet_names:
        raise ValueError("Sheet named 'Report' not found in the Excel file. Please rename the sheet you want formatted to 'Report' and try again")

    # Read the "Report" sheet
    df = xls.parse("Report")


    
    # Remove column "TOTAL" (column B)
    df = df.drop(columns=["TOTAL"])
    
    #Rename the column
    df.rename(columns={"Supplier / County": "Supplier"}, inplace=True)


    df_formatted = pd.melt(df, id_vars=["Supplier"], var_name="County", value_name="Status")

    # Change "Once" values to "Yes" or "No"
    df_formatted["Status"] = df_formatted["Status"].apply(lambda x: "Yes" if x == 1 else "No" if pd.isna(x) else x)

    return df_formatted

#=======================================================================================================================
# END function takes the Supplier by county pivot table and formats it so it can be uploaded to snowflake
#=======================================================================================================================

#-----------------------------------------------------------------------------------------------------------------------

#=======================================================================================================================
# This function will write the Supplier County to a snowflake table for gap report county validation
#=======================================================================================================================
def write_to_snowflake(df_content):
    try:
        # Get connection object
        conn, connection_id = create_snowflake_connection()
        
        # Create a cursor object to execute SQL queries
        cursor = conn.cursor()

        # Truncate the existing table before inserting new data (optional)
        truncate_query = "TRUNCATE TABLE SUPPLIER_COUNTY"
        cursor.execute(truncate_query)

        # Create an INSERT INTO query to insert the DataFrame data into the table
        insert_query = """
        INSERT INTO SUPPLIER_COUNTY (
            SUPPLIER,
            COUNTY,
            STATUS
        )
        VALUES (%s, %s, %s)
        """

        # Flatten the DataFrame and convert to list of tuples
        values = df_content[["Supplier", "County", "Status"]].values.tolist()

        # Execute the insert query
        cursor.executemany(insert_query, values)

        # Commit the changes to the Snowflake table
        conn.commit()

        # Close the connection
        conn.close()

        st.success("Data written to Snowflake successfully!")

    except snowflake.connector.errors.Error as e:
        st.error(f"Error writing to Snowflake: {str(e)}")
        
    



    
#=================================================================================================================
# End of function will write the Supplier County to a snowflake table for gap report county validation
#=================================================================================================================

#-----------------------------------------------------------------------------------------------------------------

#====================================================================================================================================
# This block of code creates an upload widget and creates the button to start the reformatting process of the Emcompess pivot table
# for supplier_county table and once file is reformatted you will be able to download it
#====================================================================================================================================


with file_container:
    st.markdown("<hr>", unsafe_allow_html=True)
    st.subheader(":blue[Process Supplier by County Formatting and Uploading to Snowflake]")
    # Create file uploader
    uploaded_file = st.file_uploader(":red[Upload Supplier by County Excel file Pivot Table from Encompess]", type=["xlsx", "xls"])

    # Initialize session state
    if "df_formatted" not in st.session_state:
        st.session_state.df_formatted = None

    # Display the formatted DataFrame
    if uploaded_file is not None:
        # Show the Reformat button
        if st.button("Reformat Supplier by County Spreadsheet", key="reformat_button"):
            # Format the Supplier BY COUNTY excel spreadsheet
            file_content = uploaded_file.getvalue()  # Get the content of the uploaded file
            df_formatted = format_supplier_by_county(file_content)
            st.session_state.df_formatted = df_formatted
            new_csv_file = 'formatted_' + uploaded_file.name
            stream = BytesIO()
            df_formatted.to_excel(stream, index=False)
            stream.seek(0)
            st.download_button(label="Download formatted Supplier by County", data=stream.read(), file_name=new_csv_file, mime='application/vnd.ms-excel')
    
#====================================================================================================================================
# End block of code creates an upload widget and creates the button to start the reformatting process of the Emcompess pivot table
# once file is reformatted you will be able to download it
#====================================================================================================================================

#------------------------------------------------------------------------------------------------------------------------------------


#====================================================================================================================================
# This code block creates and uploader for the formatted file, creates the upload button to start the file upload to snowflake
# for the Supplier by County data
#====================================================================================================================================
              


uploaded_file_snowflake = st.file_uploader(":red[Upload formatted Supplier by County file to write to Snowflake]", type=["xlsx", "xls"])

if uploaded_file_snowflake is not None:
    # Show the Write to Snowflake button
    if st.button("Write Supplier by County Data to Snowflake", key="supplier_county_button"):
        # Read the content of the uploaded file into a DataFrame
        df_uploaded = pd.read_excel(uploaded_file_snowflake)

        # Write the DataFrame data to Snowflake table
        with st.spinner('Uploading data to Snowflake ...'):
            write_to_snowflake(df_uploaded)
            st.success("Data written to Snowflake successfully!")
else:
    st.warning("Please upload the formatted file to write Supplier by County Data to Snowflake'.")
# Add horizontal line
st.markdown("<hr>", unsafe_allow_html=True)



#====================================================================================================================================
# END code block creates and uploader for the formatted file, creates the upload button to start the file upload to snowflake
# for the Supplier by County data
#====================================================================================================================================